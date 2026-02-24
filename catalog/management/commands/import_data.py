import os
import shutil
from datetime import datetime
from django.core.management.base import BaseCommand, CommandError

try:
    import openpyxl
except ImportError:
    openpyxl = None


class Command(BaseCommand):
    help = 'Импорт данных из Excel-файлов'

    def add_arguments(self, parser):
        parser.add_argument('--path', type=str, default='.', help='Путь к папке с Excel-файлами')
        parser.add_argument('--images-path', type=str, default='.', help='Путь к папке с изображениями')

    def handle(self, *args, **options):
        if openpyxl is None:
            raise CommandError('Установите openpyxl: pip install openpyxl')

        from catalog.models import Role, User, Category, Manufacturer, Supplier, Product, DeliveryPoint, Order

        path = options['path']
        images_path = options.get('images_path', path)

        self.stdout.write('=== Импорт данных ===')
        self._create_roles()

        dp_file = os.path.join(path, 'Пункты выдачи_import.xlsx')
        if os.path.exists(dp_file):
            self._import_delivery_points(dp_file)

        products_file = os.path.join(path, 'Tovar.xlsx')
        if os.path.exists(products_file):
            self._import_products(products_file, images_path)

        users_file = os.path.join(path, 'user_import.xlsx')
        if os.path.exists(users_file):
            self._import_users(users_file)

        orders_file = os.path.join(path, 'Заказ_import.xlsx')
        if os.path.exists(orders_file):
            self._import_orders(orders_file)

        self.stdout.write(self.style.SUCCESS('Импорт завершён!'))

    def _create_roles(self):
        from catalog.models import Role
        for role_name in [Role.CLIENT, Role.MANAGER, Role.ADMIN]:
            role, created = Role.objects.get_or_create(name=role_name)
            if created:
                self.stdout.write(f'  + Роль: {role.get_name_display()}')

    def _import_delivery_points(self, filepath):
        from catalog.models import DeliveryPoint
        wb = openpyxl.load_workbook(filepath)
        ws = wb.active
        count = 0
        for row in ws.iter_rows(min_row=1, values_only=True):
            if row[0]:
                address = str(row[0]).strip()
                if address:
                    DeliveryPoint.objects.get_or_create(address=address)
                    count += 1
        self.stdout.write(f'  Пункты выдачи: {count}')

    def _import_products(self, filepath, images_path):
        from catalog.models import Category, Manufacturer, Supplier, Product
        from django.conf import settings

        wb = openpyxl.load_workbook(filepath)
        ws = wb.active
        headers = [str(c.value).strip() if c.value else '' for c in ws[1]]
        count = 0

        for row in ws.iter_rows(min_row=2, values_only=True):
            if not row[0]:
                continue
            data = dict(zip(headers, row))

            category, _ = Category.objects.get_or_create(
                name=str(data.get('Категория товара', 'Без категории')).strip()
            )
            manufacturer, _ = Manufacturer.objects.get_or_create(
                name=str(data.get('Производитель', 'Неизвестен')).strip()
            )
            supplier, _ = Supplier.objects.get_or_create(
                name=str(data.get('Поставщик', 'Неизвестен')).strip()
            )

            try:
                discount = float(str(data.get('Действующая скидка', 0) or 0).replace('%', '').strip())
            except (ValueError, TypeError):
                discount = 0.0

            try:
                price = float(str(data.get('Цена', 0)).replace(',', '.'))
            except (ValueError, TypeError):
                price = 0.0

            try:
                stock = int(data.get('Кол-во на складе', 0) or 0)
            except (ValueError, TypeError):
                stock = 0

            article = str(data.get('Артикул', '')).strip()

            product, _ = Product.objects.update_or_create(
                article=article,
                defaults={
                    'name': str(data.get('Наименование товара', '')).strip(),
                    'unit': str(data.get('Единица измерения', 'пара')).strip(),
                    'price': price,
                    'discount': discount,
                    'stock': stock,
                    'description': str(data.get('Описание товара', '') or '').strip(),
                    'category': category,
                    'manufacturer': manufacturer,
                    'supplier': supplier,
                }
            )

            photo = str(data.get('Фото', '') or '').strip()
            if photo and not product.image:
                src = os.path.join(images_path, photo)
                if os.path.exists(src):
                    dest_dir = os.path.join(settings.MEDIA_ROOT, 'products')
                    os.makedirs(dest_dir, exist_ok=True)
                    shutil.copy2(src, os.path.join(dest_dir, photo))
                    product.image = f'products/{photo}'
                    product.save()

            count += 1

        self.stdout.write(f'  Товары: {count}')

    def _import_users(self, filepath):
        from catalog.models import Role, User
        from django.contrib.auth.hashers import make_password

        wb = openpyxl.load_workbook(filepath)
        ws = wb.active
        headers = [str(c.value).strip() if c.value else '' for c in ws[1]]

        role_map = {
            'Администратор': Role.ADMIN,
            'Менеджер': Role.MANAGER,
            'Авторизованный клиент': Role.CLIENT,
        }
        count = 0

        for row in ws.iter_rows(min_row=2, values_only=True):
            if not row[0]:
                continue
            data = dict(zip(headers, row))
            role_key = role_map.get(str(data.get('Роль сотрудника', '')).strip(), Role.CLIENT)
            try:
                role = Role.objects.get(name=role_key)
            except Role.DoesNotExist:
                continue

            username = str(data.get('Логин', '')).strip()
            if not username:
                continue

            User.objects.get_or_create(
                username=username,
                defaults={
                    'full_name': str(data.get('ФИО', '')).strip(),
                    'role': role,
                    'password': make_password(str(data.get('Пароль', '')).strip()),
                }
            )
            count += 1

        self.stdout.write(f'  Пользователи: {count}')

    def _import_orders(self, filepath):
        from catalog.models import Order, DeliveryPoint

        wb = openpyxl.load_workbook(filepath)
        ws = wb.active
        headers = [str(c.value).strip() if c.value else '' for c in ws[1]]

        status_map = {
            'Завершен': Order.STATUS_COMPLETED,
            'Новый': Order.STATUS_NEW,
            'Отменен': Order.STATUS_CANCELLED,
        }
        count = 0

        for row in ws.iter_rows(min_row=2, values_only=True):
            if not row[0]:
                continue
            data = dict(zip(headers, row))

            try:
                number = int(data.get('Номер заказа', 0))
            except (ValueError, TypeError):
                continue

            dp_address = str(data.get('Адрес пункта выдачи', '') or '').strip()
            delivery_point = DeliveryPoint.objects.filter(address__icontains=dp_address[:30]).first() if dp_address else None

            def parse_date(val, label):
                if isinstance(val, str):
                    try:
                        return datetime.strptime(val, '%d.%m.%Y').date()
                    except ValueError:
                        self.stdout.write(self.style.WARNING(
                            f'  ⚠ Заказ №{number}: невалидная дата {label} "{val}" — заменена на сегодняшнюю.'
                        ))
                        return datetime.now().date()
                elif hasattr(val, 'date'):
                    return val.date()
                return None

            order_date = parse_date(data.get('Дата заказа'), 'заказа') or datetime.now().date()
            delivery_date = parse_date(data.get('Дата доставки'), 'доставки')

            Order.objects.get_or_create(
                number=number,
                defaults={
                    'article': str(data.get('Артикул заказа', '') or '').strip(),
                    'order_date': order_date,
                    'delivery_date': delivery_date,
                    'delivery_point': delivery_point,
                    'client_name': str(data.get('ФИО авторизированного клиента', '') or '').strip(),
                    'pickup_code': str(data.get('Код для получения', '') or '').strip(),
                    'status': status_map.get(str(data.get('Статус заказа', 'Новый')).strip(), Order.STATUS_NEW),
                }
            )
            count += 1

        self.stdout.write(f'  Заказы: {count}')
